#!/usr/bin/env python3
"""
A-B 区段运行图指标计算 — 附表1 & 附表2
==========================================
基于题目 4.11 数据，按标准表格格式计算并输出：
  附表1：区间通过能力计算表
  附表2：运行图质量指标计算表

用法: python calc_tables.py
"""

import sys, os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 修复 Windows 控制台 GBK 编码
try:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

# 从 train_diagram 导入基础数据
from train_diagram import (
    STATIONS, N_STATIONS, DISTANCES, TOTAL_DISTANCE,
    DOWN_RUNNING, UP_RUNNING,
    T_START, T_STOP, T_QITING,
    TAU_UNSIM, TAU_MEET,
    T_MAINTENANCE,
    PASSENGER_TRAINS, N_PASSENGER,
    FREIGHT_DOWN, FREIGHT_UP,
    EPSILON_PASSENGER, EPSILON_FAST_FREIGHT, EPSILON_PICKUP,
    N_LOCOMOTIVES,
    A_LOCO_TURNAROUND, B_LOCO_TURNAROUND,
    B_STATION_OPERATION_TIME,
    STOP_PASSENGER, STOP_FREIGHT, STOP_PICKUP,
    calc_train_schedule, calc_restrictive_section,
)


# ================================================================
# 第一部分：附表1 — A-B 区段区间通过能力计算表
# ================================================================

def get_meet_plan(section_idx):
    """
    根据两端车站股道数选择会车方案（4种标准方案）：
      (a) 两端不停车通过，进入区间 — 两端均有≥3股道
      (b) 两端不停车通过，开出区间 — 两端均有≥3股道
      (c) 下行不停车通过 — 上行端站股道少，下行方向优先通过
      (d) 上行不停车通过 — 下行端站股道少，上行方向优先通过
    """
    from train_diagram import TRACKS
    track_a = TRACKS[section_idx]       # 区间左端站
    track_b = TRACKS[section_idx + 1]   # 区间右端站

    if section_idx == 0:  # A-a: A=5, a=3 → 两端大站
        return "(a)两端不停车通过"
    elif section_idx == 1:  # a-b: a=3, b=2 → b站只有2股道
        return "(c)下行不停车通过"
    elif section_idx == 2:  # b-c: b=2, c=3 → b站只有2股道
        return "(d)上行不停车通过"
    elif section_idx == 3:  # c-d: c=3, d=3 → 两端OK
        return "(a)两端不停车通过"
    elif section_idx == 4:  # d-e: d=3, e=2 → e站只有2股道
        return "(c)下行不停车通过"
    else:  # e-B: e=2, B=5 → e站只有2股道
        return "(d)上行不停车通过"


def print_table1(parallel_result, non_parallel_result):
    """打印附表1：A-B 区段区间通过能力计算表 (按站排列，τ为车站属性)"""

    sections = parallel_result["section_details"]
    n_parallel = parallel_result["n_parallel"]

    print("\n" + "=" * 140)
    print("  附表1  A-B 区段区间通过能力计算表")
    print("=" * 140)

    # 表头
    header1 = (
        f"{'':─^6} {'':─^8} {'':─^10} {'':─^8} {'区间运行时间(min)':─^34} "
        f"{'车站间隔时间(min)':─^26} {'':─^10} {'':─^8} {'':─^8} {'T周':─^8} {'N平行':─^8}"
    )
    header2 = (
        f"{'站名':^6} {'到发线':^8} {'闭塞方法':^10} {'区间距离':^8} "
        f"{'上行':^8} {'下行':^8} {'t起':^8} {'t停':^8} "
        f"{'τ不':^8} {'τ会':^8} {'τ空1':^8} {'τ空2':^8} "
        f"{'t技停':^8} {'会车方案':^8} {'(min)':^8} {'(对/天)':^8}"
    )
    print(header1)
    print(header2)
    print("─" * 140)

    # τ不/τ会 按站的值 (7个站)
    # A: τ会=2, τ不=-  |  a: τ不=5 τ会=3  |  b: τ不=5 τ会=3
    # c: τ不=5 τ会=3  |  d: τ不=5 τ会=3  |  e: τ不=5 τ会=3  |  B: τ会=2 τ不=-
    TAU_BU_STATION  = [0, 5, 5, 5, 5, 5, 0]   # τ不 by station
    TAU_HUI_STATION = [2, 3, 3, 3, 3, 3, 2]   # τ会 by station
    TRACKS_STATION  = [5, 3, 2, 3, 3, 2, 5]   # 到发线
    T_START_STATION = [0, 1, 1, 1, 1, 1, 0]   # t起
    T_STOP_STATION  = [0, 1, 1, 1, 1, 1, 0]   # t停

    # 运行时分 (6个区间)
    from train_diagram import UP_RUNNING, DOWN_RUNNING, DISTANCES

    available_time = 1440 - T_MAINTENANCE

    # 逐站输出 (7站)
    for station_idx in range(N_STATIONS):
        station = STATIONS[station_idx]
        track_count = TRACKS_STATION[station_idx]
        tau_b = TAU_BU_STATION[station_idx]
        tau_h = TAU_HUI_STATION[station_idx]
        t_start = T_START_STATION[station_idx]
        t_stop = T_STOP_STATION[station_idx]

        # 区间数据 (本站到下一站, 终端站B无下一区间)
        if station_idx < N_STATIONS - 1:
            sec = sections[station_idx]
            dist = DISTANCES[station_idx]
            t_up = UP_RUNNING[station_idx]
            t_down = DOWN_RUNNING[station_idx]
            t_period = sec["t_period"]
            n_par_sec = available_time / t_period
            meet = get_meet_plan(station_idx)
        else:
            dist = "—"
            t_up = "—"
            t_down = "—"
            t_period = "—"
            n_par_sec = "—"
            meet = "—"

        t_up_str = f"{t_up:^8}" if isinstance(t_up, int) else f"{t_up:^8}"
        t_down_str = f"{t_down:^8}" if isinstance(t_down, int) else f"{t_down:^8}"
        t_per_str = f"{t_period:^8}" if isinstance(t_period, int) else f"{t_period:^8}"
        n_par_str = f"{n_par_sec:^8.1f}" if isinstance(n_par_sec, float) else f"{n_par_sec:^8}"
        dist_str = f"{dist:^8}" if isinstance(dist, int) else f"{dist:^8}"

        tau_b_str = f"{tau_b:^8}" if tau_b > 0 else f"{'—':^8}"
        tau_h_str = f"{tau_h:^8}" if tau_h > 0 else f"{'—':^8}"

        print(
            f"{station:^6} {track_count:^8} {'半自动':^10} {dist_str} "
            f"{t_up_str} {t_down_str} {t_start:^8} {t_stop:^8} "
            f"{tau_b_str} {tau_h_str} {'—':^8} {'—':^8} "
            f"{0:^8} {meet:^8} {t_per_str} {n_par_str}"
        )

    # 限制区间
    restrictive = max(sections, key=lambda s: s["t_period"])
    print("─" * 140)
    print(f"  限制区间: {restrictive['section']} (T周 = {restrictive['t_period']} min)")
    print(f"  可用时间: 1440 - {T_MAINTENANCE} = {1440 - T_MAINTENANCE} min")
    print(f"  平行运行图通过能力 N平行 = {1440 - T_MAINTENANCE} / {restrictive['t_period']} = {n_parallel} 对/天")

    # 补充：非平行运行图通过能力
    print(f"\n  非平行运行图通过能力计算:")
    print(f"  旅客列车扣除: ε客 × n客 = {EPSILON_PASSENGER} × {N_PASSENGER} = {EPSILON_PASSENGER * N_PASSENGER}")
    n_pickup = FREIGHT_DOWN["摘挂"] + FREIGHT_UP["摘挂"]
    n_fast = FREIGHT_DOWN["直达(空)"] + FREIGHT_UP["直达(重)"]
    print(f"  摘挂列车追加扣除: (ε摘挂-1) × n摘挂 = {EPSILON_PICKUP - 1} × {n_pickup} = {(EPSILON_PICKUP - 1) * n_pickup:.1f}")
    print(f"  快货列车追加扣除: (ε快货-1) × n快货 = {EPSILON_FAST_FREIGHT - 1:.1f} × {n_fast} = {(EPSILON_FAST_FREIGHT - 1) * n_fast:.1f}")
    total_deduction = EPSILON_PASSENGER * N_PASSENGER + (EPSILON_PICKUP - 1) * n_pickup + (EPSILON_FAST_FREIGHT - 1) * n_fast
    n_non_parallel = n_parallel - total_deduction
    print(f"  总扣除: {total_deduction:.1f} 对")
    print(f"  N非平行 = {n_parallel} - {total_deduction:.1f} = {n_non_parallel:.1f} → 取整 {int(n_non_parallel)} 对/天")

    return sections, n_parallel, n_non_parallel


# ================================================================
# 第二部分：附表2 — A-B 区段运行图质量指标计算表
# ================================================================

def calc_table2_data():
    """计算附表2需要的全部数据"""

    # ---------- 构建全部列车时刻 ----------
    # 旅客列车
    passenger_list = []
    for pt in PASSENGER_TRAINS:
        passenger_list.append({
            "id": pt["id"], "type": "旅客", "dir": pt["dir"],
            "depart_time": pt["depart"],
        })

    # 货物列车（下行）
    freight_list = []
    down_count = 0
    down_ids = []
    # 摘挂
    pickup_down_times = [6*60+30, 14*60+30]
    for i in range(FREIGHT_DOWN["摘挂"]):
        t = pickup_down_times[i % len(pickup_down_times)]
        fid = f"4003{5+i}"
        freight_list.append({"id": fid, "type": "摘挂", "dir": "down", "depart_time": t})
        down_ids.append(fid)
        down_count += 1
    # 区段
    for i in range(FREIGHT_DOWN["区段"]):
        t = T_MAINTENANCE + 10 + i * 130
        fid = f"3000{i+1}"
        freight_list.append({"id": fid, "type": "区段", "dir": "down", "depart_time": int(t)})
        down_ids.append(fid)
        down_count += 1
    # 直达
    for i in range(FREIGHT_DOWN["直达(空)"]):
        t = T_MAINTENANCE + 10 + (FREIGHT_DOWN["区段"] + i) * 130 + 65
        fid = f"8600{i+1}"
        freight_list.append({"id": fid, "type": "直达", "dir": "down", "depart_time": int(t)})
        down_ids.append(fid)
        down_count += 1

    # 货物列车（上行）
    up_ids = []
    # 摘挂
    pickup_up_times = [10*60+30, 18*60+30]
    for i in range(FREIGHT_UP["摘挂"]):
        t = pickup_up_times[i % len(pickup_up_times)]
        fid = f"4004{5+i}"
        freight_list.append({"id": fid, "type": "摘挂", "dir": "up", "depart_time": t})
        up_ids.append(fid)
    # 区段
    for i in range(FREIGHT_UP["区段"]):
        t = T_MAINTENANCE + 10 + i * 130 + 65
        fid = f"1000{i+2}"
        freight_list.append({"id": fid, "type": "区段", "dir": "up", "depart_time": int(t)})
        up_ids.append(fid)
    # 直达
    for i in range(FREIGHT_UP["直达(重)"]):
        t = T_MAINTENANCE + 10 + (FREIGHT_UP["区段"] + i) * 130
        fid = f"1000{i+12}"
        freight_list.append({"id": fid, "type": "直达", "dir": "up", "depart_time": int(t)})
        up_ids.append(fid)

    # ---------- 为每列车计算时刻 ----------
    all_trains_timetable = []
    for train in passenger_list + freight_list:
        schedule = calc_train_schedule(train)
        all_trains_timetable.append(schedule)

    # ---------- 按方向分组 ----------
    down_trains = []
    up_trains = []
    for t in all_trains_timetable:
        if t["dir"] == "down" and t["type"] != "旅客":
            down_trains.append(t)
        elif t["dir"] == "up" and t["type"] != "旅客":
            up_trains.append(t)

    # 旅客列车单独列出（不填附表2主表，通常旅客列车不在货物列车指标表里）
    # 但附表2要求填所有列车，这里按上下行分别列出

    # ---------- 构建附表2行数据 ----------
    rows = []

    # 对每个下行货物列车，计算指标
    for dt in down_trains:
        stations = dt["stations"]
        # 由A发
        depart_a = stations[0]["depart"]  # A站出发
        # 到n (这里的n是下行方向终到站B)
        arrive_b = stations[-1]["arrive"]  # B站到达
        # 在途时间
        if depart_a is not None and arrive_b is not None:
            travel_time = arrive_b - depart_a
            run_time = sum(DOWN_RUNNING)
            stop_time = travel_time - run_time
            train_km = TOTAL_DISTANCE
        else:
            travel_time = 0
            run_time = 0
            stop_time = 0
            train_km = 0

        rows.append({
            "down_id": dt["id"],
            "down_depart_a": depart_a,
            "down_arrive_b": arrive_b,
            "down_travel": travel_time,
            "down_run": run_time,
            "down_stop": stop_time,
            "down_km": train_km,
            "up_id": None, "up_depart_b": None, "up_arrive_a": None,
            "up_travel": None, "up_run": None, "up_stop": None, "up_km": None,
            "loco_b_stay": None, "loco_tow": None, "loco_a_depart": None, "loco_a_stay": None,
        })

    # 对每个上行货物列车，计算指标
    for ut in up_trains:
        stations = ut["stations"]
        depart_b = stations[-1]["depart"]
        arrive_a = stations[0]["arrive"]
        if depart_b is not None and arrive_a is not None:
            travel_time = arrive_a - depart_b
            run_time = sum(UP_RUNNING)
            stop_time = travel_time - run_time
            train_km = TOTAL_DISTANCE
        else:
            travel_time = 0
            run_time = 0
            stop_time = 0
            train_km = 0

        rows.append({
            "down_id": None, "down_depart_a": None, "down_arrive_b": None,
            "down_travel": None, "down_run": None, "down_stop": None, "down_km": None,
            "up_id": ut["id"],
            "up_depart_b": depart_b,
            "up_arrive_a": arrive_a,
            "up_travel": travel_time,
            "up_run": run_time,
            "up_stop": stop_time,
            "up_km": train_km,
            "loco_b_stay": None, "loco_tow": None, "loco_a_depart": None, "loco_a_stay": None,
        })

    # ---------- 机车交路计算 ----------
    # 分开下行和上行原始行
    down_rows = [r for r in rows if r["down_id"] is not None]
    up_rows = [r for r in rows if r["up_id"] is not None]

    # 按时序排序
    down_rows.sort(key=lambda r: r["down_depart_a"] or 0)
    up_rows.sort(key=lambda r: r["up_depart_b"] or 0)

    # 机车交路链：下行i→B站→上行配对→A站→下一个下行i+1
    # Step 1: 下行→上行配对 (B站折返)
    # 每个下行到达B后，找第一个满足折返时间≥90min的上行
    loco_chain = []  # [(down_row, up_row, b_stay)]
    up_used = set()
    for dr in down_rows:
        best = None
        best_stay = None
        for ur in up_rows:
            if ur["up_id"] in up_used:
                continue
            if dr["down_arrive_b"] and ur["up_depart_b"]:
                stay = ur["up_depart_b"] - dr["down_arrive_b"]
                if stay >= B_LOCO_TURNAROUND:  # ≥1.5h
                    best = ur
                    best_stay = stay
                    up_used.add(ur["up_id"])
                    break
        loco_chain.append((dr, best, best_stay))

    # 剩余未配对上行
    for ur in up_rows:
        if ur["up_id"] not in up_used:
            loco_chain.append((None, ur, None))

    # Step 2: 上行→下一行配对 (A站折返)
    # 为每个配了对的上行，找下一个下行
    a_stays = {}  # up_id -> (down_row, a_stay)
    for i, (dr, ur, b_stay) in enumerate(loco_chain):
        if ur is None:
            continue
        # 找此上行之后的下一个下行（从A出发晚于上行到达A，且满足折返时间≥120min）
        for j in range(i + 1, len(loco_chain)):
            next_dr = loco_chain[j][0]
            if next_dr and ur["up_arrive_a"] and next_dr["down_depart_a"]:
                stay = next_dr["down_depart_a"] - ur["up_arrive_a"]
                if stay >= A_LOCO_TURNAROUND:  # ≥2.0h
                    a_stays[ur["up_id"]] = (next_dr["down_id"], next_dr["down_depart_a"], stay)
                    break

    # Step 3: 合并成最终行
    final_rows = []
    for dr, ur, b_stay in loco_chain:
        row = {
            "down_id": dr["down_id"] if dr else None,
            "down_depart_a": dr["down_depart_a"] if dr else None,
            "down_arrive_b": dr["down_arrive_b"] if dr else None,
            "down_travel": dr["down_travel"] if dr else None,
            "down_run": dr["down_run"] if dr else None,
            "down_stop": dr["down_stop"] if dr else None,
            "down_km": dr["down_km"] if dr else None,
            "up_id": ur["up_id"] if ur else None,
            "up_depart_b": ur["up_depart_b"] if ur else None,
            "up_arrive_a": ur["up_arrive_a"] if ur else None,
            "up_travel": ur["up_travel"] if ur else None,
            "up_run": ur["up_run"] if ur else None,
            "up_stop": ur["up_stop"] if ur else None,
            "up_km": ur["up_km"] if ur else None,
            "loco_b_stay": b_stay,
            "loco_tow": ur["up_id"] if ur else None,
            "loco_a_depart": None,
            "loco_a_stay": None,
        }
        # 填入A站折返数据
        if ur and ur["up_id"] in a_stays:
            row["loco_a_depart"] = a_stays[ur["up_id"]][1]
            row["loco_a_stay"] = a_stays[ur["up_id"]][2]
        final_rows.append(row)

    return final_rows


def print_table2(rows):
    """打印附表2：A-B 区段运行图质量指标计算表"""

    print("\n" + "=" * 160)
    print("  附表2  A-B 区段运行图质量指标计算表")
    print("=" * 160)

    # 表头
    header = (
        f"{'下行方向':─^55} {'上行方向':─^55} {'机车':─^42}"
    )
    sub_header = (
        f"{'车次':^8} {'由A发':^8} {'到n':^8} {'在途':^8} {'运行':^8} {'停站':^8} {'列公里':^8} "
        f"{'车次':^8} {'由B发':^8} {'到A':^8} {'在途':^8} {'运行':^8} {'停站':^8} {'列公里':^8} "
        f"{'B站停留':^10} {'牵引车次':^10} {'A站出发':^10} {'A站停留':^10}"
    )
    print(header)
    print(sub_header)
    print("─" * 160)

    # 数据行
    total_down_run = 0
    total_down_stop = 0
    total_down_km = 0
    total_up_run = 0
    total_up_stop = 0
    total_up_km = 0

    for row in rows:
        def fmt_time(t):
            if t is None:
                return "  —"
            t = int(t)
            return f"{t//60:02d}:{t%60:02d}"

        def fmt_min(v):
            if v is None:
                return "  —"
            return f"{int(v):^8}"

        def fmt_km(v):
            if v is None:
                return "  —"
            return f"{int(v):^8}"

        # 下行
        d_id = row["down_id"] or "—"
        d_dep = fmt_time(row["down_depart_a"])
        d_arr = fmt_time(row["down_arrive_b"])
        d_travel = fmt_min(row["down_travel"])
        d_run = fmt_min(row["down_run"])
        d_stop = fmt_min(row["down_stop"])
        d_km = fmt_km(row["down_km"])

        # 上行
        u_id = row["up_id"] or "—"
        u_dep = fmt_time(row["up_depart_b"])
        u_arr = fmt_time(row["up_arrive_a"])
        u_travel = fmt_min(row["up_travel"])
        u_run = fmt_min(row["up_run"])
        u_stop = fmt_min(row["up_stop"])
        u_km = fmt_km(row["up_km"])

        # 机车
        b_stay = fmt_min(row["loco_b_stay"])
        tow = row["loco_tow"] or "—"
        a_dep = fmt_time(row["loco_a_depart"])
        a_stay = fmt_min(row["loco_a_stay"])

        print(
            f"{d_id:^8} {d_dep:^8} {d_arr:^8} {d_travel:^8} {d_run:^8} {d_stop:^8} {d_km:^8} "
            f"{u_id:^8} {u_dep:^8} {u_arr:^8} {u_travel:^8} {u_run:^8} {u_stop:^8} {u_km:^8} "
            f"{b_stay:^10} {tow:^10} {a_dep:^10} {a_stay:^10}"
        )

        if row["down_run"]:
            total_down_run += row["down_run"]
            total_down_stop += row["down_stop"]
            total_down_km += row["down_km"]
        if row["up_run"]:
            total_up_run += row["up_run"]
            total_up_stop += row["up_stop"]
            total_up_km += row["up_km"]

    # 合计行
    print("─" * 160)
    print(
        f"{'合计':^8} {'—':^8} {'—':^8} {str(total_down_run+total_down_stop):^8} "
        f"{str(total_down_run):^8} {str(total_down_stop):^8} {str(total_down_km):^8} "
        f"{'合计':^8} {'—':^8} {'—':^8} {str(total_up_run+total_up_stop):^8} "
        f"{str(total_up_run):^8} {str(total_up_stop):^8} {str(total_up_km):^8} "
        f"{'—':^10} {'—':^10} {'—':^10} {'—':^10}"
    )

    # 速度指标
    n_down = len([r for r in rows if r["down_id"] is not None])
    n_up = len([r for r in rows if r["up_id"] is not None])

    if n_down > 0 and total_down_run > 0:
        tech_speed_d = (total_down_km / total_down_run) * 60
        travel_speed_d = (total_down_km / (total_down_run + total_down_stop)) * 60
        speed_coef_d = travel_speed_d / tech_speed_d
    else:
        tech_speed_d = travel_speed_d = speed_coef_d = 0

    if n_up > 0 and total_up_run > 0:
        tech_speed_u = (total_up_km / total_up_run) * 60
        travel_speed_u = (total_up_km / (total_up_run + total_up_stop)) * 60
        speed_coef_u = travel_speed_u / tech_speed_u
    else:
        tech_speed_u = travel_speed_u = speed_coef_u = 0

    total_km = total_down_km + total_up_km
    total_run = total_down_run + total_up_run
    total_travel = total_down_run + total_down_stop + total_up_run + total_up_stop

    print(f"\n  注：每一行上行方向车次按机车折返交路的顺序填写")
    print(f"\n  汇总指标:")
    print(f"  {'':<20} {'下行':<15} {'上行':<15} {'合计':<15}")
    print(f"  {'─' * 60}")
    print(f"  {'统计列车数':<20} {n_down:<15} {n_up:<15} {n_down+n_up:<15}")
    print(f"  {'技术速度(km/h)':<20} {tech_speed_d:<15.1f} {tech_speed_u:<15.1f} {(total_km/total_run*60) if total_run else 0:<15.1f}")
    print(f"  {'旅行速度(km/h)':<20} {travel_speed_d:<15.1f} {travel_speed_u:<15.1f} {(total_km/total_travel*60) if total_travel else 0:<15.1f}")
    print(f"  {'速度系数':<20} {speed_coef_d:<15.4f} {speed_coef_u:<15.4f} {((total_km/total_travel*60)/(total_km/total_run*60)) if total_run and total_travel else 0:<15.4f}")

    # 机车指标
    total_b_stay = sum(r["loco_b_stay"] for r in rows if r["loco_b_stay"])
    total_a_stay = sum(r["loco_a_stay"] for r in rows if r["loco_a_stay"])
    n_cycles = len([r for r in rows if r["loco_b_stay"]])

    print(f"\n  机车运用指标:")
    print(f"  {'机车交路数':<20} {n_cycles}")
    print(f"  {'平均B站停留(min)':<20} {total_b_stay/n_cycles if n_cycles else 0:.0f}")
    print(f"  {'平均A站停留(min)':<20} {total_a_stay/n_cycles if n_cycles else 0:.0f}")

    if n_cycles > 0:
        avg_cycle = (total_travel / (n_down + n_up)) * 2 + (total_b_stay + total_a_stay) / n_cycles
        print(f"  {'平均周转时间(min)':<20} {avg_cycle:.0f}")
        daily_km_total = (1440 / avg_cycle) * TOTAL_DISTANCE * 2 * N_LOCOMOTIVES
        daily_km_per = (1440 / avg_cycle) * TOTAL_DISTANCE * 2
        print(f"  {'机车日车公里(单台)':<20} {daily_km_per:.0f} km")
        print(f"  {'机车日车公里(总)':<20} {daily_km_total:.0f} km")


# ================================================================
# 第三部分：Excel 输出
# ================================================================

# 通用样式
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FONT = Font(name="微软雅黑", bold=True, size=10)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_W = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
TITLE_FONT = Font(name="微软雅黑", bold=True, size=14)
DATA_FONT = Font(name="微软雅黑", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header_row(ws, row, max_col, fill=None, font=None):
    """给表头行设置样式"""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        if fill:
            cell.fill = fill
        if font:
            cell.font = font


def style_data_row(ws, row, max_col):
    """给数据行设置样式"""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        cell.font = DATA_FONT


def auto_width(ws, max_col, min_width=8, max_width=18):
    """自动列宽"""
    for col in range(1, max_col + 1):
        max_len = min_width
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            for val in row:
                if val:
                    max_len = max(max_len, min(len(str(val)) * 1.3, max_width))
        ws.column_dimensions[get_column_letter(col)].width = max_len


def save_to_xlsx(parallel_result, non_parallel_result, table2_rows):
    """将附表1和附表2写入xlsx文件 — 按题目原表23列/18列精确格式"""

    wb = openpyxl.Workbook()

    # ================================================================
    # Sheet 1: 附表1 — 23列 × (6区间 + 限制分析 + 非平行扣除)
    # ================================================================
    ws1 = wb.active
    ws1.title = "附表1-通过能力"
    N_COL1 = 23

    sections = parallel_result["section_details"]
    n_parallel = parallel_result["n_parallel"]
    n_pickup = FREIGHT_DOWN["摘挂"] + FREIGHT_UP["摘挂"]
    n_fast = FREIGHT_DOWN["直达(空)"] + FREIGHT_UP["直达(重)"]
    total_ded = EPSILON_PASSENGER * N_PASSENGER + (EPSILON_PICKUP - 1) * n_pickup + (EPSILON_FAST_FREIGHT - 1) * n_fast
    n_non = int(n_parallel - total_ded)

    # -- 标题 --
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N_COL1)
    c = ws1.cell(row=1, column=1, value="A-B 区段区间通过能力计算表         附表 1")
    c.font = TITLE_FONT; c.alignment = CENTER

    # -- 表头 Row 2 (第1层分类) --
    r2_labels = {
        1: "站名", 2: "到发线", 3: "闭塞方法", 4: "区间距离",
    }
    # 合并区间运行时间 (5-8), 车站间隔时间 (9-12)
    merges_r2 = [
        (5, 8, "区间运行时间"), (9, 12, "车站间隔时间\n（分）"),
        (17, 18, "旅客列车"), (19, 20, "摘挂列车"),
    ]
    singles_r2 = {
        13: "技术作业\n停站时间", 14: "会车方案", 15: "T周", 16: "平行图\n最大能力",
        21: "n非货", 22: "n非", 23: "附记",
    }

    for col, label in r2_labels.items():
        ws1.cell(row=2, column=col, value=label)
    for sc, ec, label in merges_r2:
        ws1.merge_cells(start_row=2, start_column=sc, end_row=2, end_column=ec)
        ws1.cell(row=2, column=sc, value=label)
    for col, label in singles_r2.items():
        ws1.cell(row=2, column=col, value=label)

    # -- 表头 Row 3 (第2层细分) --
    r3_data = {
        1: " ", 2: " ", 3: " ", 4: " ",
        5: "上", 6: "下", 7: "t起", 8: "t停",
        9: "τ不", 10: "τ会", 11: " ", 12: " ",
        13: " ", 14: " ", 15: " ", 16: " ",
        17: "ε客", 18: "n客", 19: "ε摘", 20: "n摘",
        21: " ", 22: " ", 23: " ",
    }
    for col, label in r3_data.items():
        ws1.cell(row=3, column=col, value=label)

    # -- 表头 Row 4 (列号 1-23) --
    for col in range(1, N_COL1 + 1):
        ws1.cell(row=4, column=col, value=col)

    # 合并纵跨行
    for col in [1, 2, 3, 4, 13, 14, 15, 16, 21, 22, 23]:
        ws1.merge_cells(start_row=2, start_column=col, end_row=4, end_column=col)

    # 样式表头
    for r in [2, 3, 4]:
        style_header_row(ws1, r, N_COL1, fill=HEADER_FILL, font=HEADER_FONT_W)

    # 站级数据 (τ不/τ会/t起/t停 是车站属性)
    TAU_BU_ST  = [0, 5, 5, 5, 5, 5, 0]   # τ不 by station
    TAU_HUI_ST = [2, 3, 3, 3, 3, 3, 2]   # τ会 by station
    TRACKS_ST  = [5, 3, 2, 3, 3, 2, 5]   # 到发线
    TS_START   = [0, 1, 1, 1, 1, 1, 0]   # t起
    TS_STOP    = [0, 1, 1, 1, 1, 1, 0]   # t停
    avail = 1440 - T_MAINTENANCE

    # -- 数据行 (7个站) --
    for st_idx in range(N_STATIONS):
        r = 5 + st_idx
        station = STATIONS[st_idx]
        tau_b = TAU_BU_ST[st_idx]
        tau_h = TAU_HUI_ST[st_idx]
        t_start = TS_START[st_idx]
        t_stop = TS_STOP[st_idx]

        if st_idx < N_STATIONS - 1:
            sec = sections[st_idx]
            dist = DISTANCES[st_idx]
            t_up = UP_RUNNING[st_idx]
            t_down = DOWN_RUNNING[st_idx]
            t_per = sec["t_period"]
            n_par = round(avail / t_per, 1)
            meet = get_meet_plan(st_idx)
        else:
            dist = "—"; t_up = "—"; t_down = "—"; t_per = "—"; n_par = "—"; meet = "—"

        data = {
            1: station, 2: TRACKS_ST[st_idx], 3: "半自动", 4: dist,
            5: t_up, 6: t_down, 7: t_start, 8: t_stop,
            9: tau_b if tau_b > 0 else "—", 10: tau_h if tau_h > 0 else "—", 11: "", 12: "",
            13: 0, 14: meet, 15: t_per, 16: n_par,
            17: EPSILON_PASSENGER, 18: N_PASSENGER,
            19: EPSILON_PICKUP, 20: n_pickup,
            21: "", 22: "", 23: "",
        }
        for col, val in data.items():
            ws1.cell(row=r, column=col, value=val)
        style_data_row(ws1, r, N_COL1)

    # 限制区间行
    r = 5 + N_STATIONS
    restrictive = max(sections, key=lambda s: s["t_period"])
    for col in range(1, N_COL1 + 1):
        ws1.cell(row=r, column=col, value="").border = THIN_BORDER
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N_COL1)
    ws1.cell(row=r, column=1,
             value=f"限制区间: {restrictive['section']}  T周={restrictive['t_period']} min  "
                   f"可用时间=1440-{T_MAINTENANCE}={1440-T_MAINTENANCE} min  "
                   f"N平行={1440-T_MAINTENANCE}/{restrictive['t_period']}={n_parallel} 对/天")
    ws1.cell(row=r, column=1).font = Font(name="微软雅黑", bold=True, size=10)
    ws1.cell(row=r, column=1).alignment = CENTER

    # 非平行扣除行
    r += 2
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N_COL1)
    ws1.cell(row=r, column=1, value="非平行运行图通过能力计算").font = Font(name="微软雅黑", bold=True, size=10)
    r += 1
    deductions = [
        f"ε客={EPSILON_PASSENGER}  n客={N_PASSENGER}  扣除: {EPSILON_PASSENGER}×{N_PASSENGER}={EPSILON_PASSENGER*N_PASSENGER:.1f}",
        f"ε摘={EPSILON_PICKUP}  n摘={n_pickup}  追加扣除: ({EPSILON_PICKUP}-1)×{n_pickup}={(EPSILON_PICKUP-1)*n_pickup:.1f}",
        f"ε快货={EPSILON_FAST_FREIGHT}  n快货={n_fast}  追加扣除: ({EPSILON_FAST_FREIGHT}-1)×{n_fast}={(EPSILON_FAST_FREIGHT-1)*n_fast:.1f}",
        f"总扣除={total_ded:.1f} 对   N非平行={n_parallel}-{total_ded:.1f}={n_parallel-total_ded:.1f}  取整={n_non} 对/天",
    ]
    for d in deductions:
        ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N_COL1)
        ws1.cell(row=r, column=1, value=d).font = DATA_FONT
        r += 1

    auto_width(ws1, N_COL1, min_width=5, max_width=14)

    # ================================================================
    # Sheet 2: 附表2 — 18列 × (车次行 + 合计 + 指标汇总)
    # ================================================================
    ws2 = wb.create_sheet("附表2-质量指标")
    N_COL2 = 18

    # -- 标题 --
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N_COL2)
    c = ws2.cell(row=1, column=1, value="A-B 区段运行图质量指标计算表         附表 2")
    c.font = TITLE_FONT; c.alignment = CENTER

    # -- 表头 Row 2 (三大区域) --
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    ws2.cell(row=2, column=1, value="下 行 方 向")
    ws2.merge_cells(start_row=2, start_column=8, end_row=2, end_column=14)
    ws2.cell(row=2, column=8, value="上 行 方 向")
    ws2.merge_cells(start_row=2, start_column=15, end_row=2, end_column=18)
    ws2.cell(row=2, column=15, value="机  车")

    # -- 表头 Row 3 (次级) --
    ws2.merge_cells(start_row=3, start_column=2, end_row=3, end_column=6)
    ws2.cell(row=3, column=2, value="时    间")
    ws2.merge_cells(start_row=3, start_column=9, end_row=3, end_column=13)
    ws2.cell(row=3, column=9, value="时    间")

    r3_singles = {
        1: "车次", 7: "列车公里",
        8: "车次", 14: "列车公里",
        15: "在B站\n停留时间", 16: "从A站\n牵引车次", 17: "从A站\n出发时间", 18: "在A站\n停留时间",
    }
    for col, label in r3_singles.items():
        ws2.cell(row=3, column=col, value=label)

    # -- 表头 Row 4 (时间细分) --
    r4_data = {
        1: " ", 2: "由A发", 3: "到n", 4: "在途",
        7: " ",
        8: " ", 9: "由B发", 10: "到A", 11: "在途",
        14: " ",
        15: " ", 16: " ", 17: " ", 18: " ",
    }
    for col, label in r4_data.items():
        ws2.cell(row=4, column=col, value=label)
    # "其中" 合并 5-6
    ws2.merge_cells(start_row=4, start_column=5, end_row=4, end_column=6)
    ws2.cell(row=4, column=5, value="其  中")
    ws2.merge_cells(start_row=4, start_column=12, end_row=4, end_column=13)
    ws2.cell(row=4, column=12, value="其  中")

    # -- 表头 Row 5 (运行/停站) --
    r5_data = {
        5: "运行", 6: "停站", 12: "运行", 13: "停站",
    }
    for col in range(1, N_COL2 + 1):
        ws2.cell(row=5, column=col, value=r5_data.get(col, " "))

    # 合并纵跨行
    for col in [1, 7, 8, 14, 15, 16, 17, 18]:
        ws2.merge_cells(start_row=3, start_column=col, end_row=5, end_column=col)

    # -- 表头 Row 6 (列号) --
    for col in range(1, N_COL2 + 1):
        ws2.cell(row=6, column=col, value=col)

    # 样式表头
    HEADER_ROW_START = 2
    for r in range(HEADER_ROW_START, 7):
        style_header_row(ws2, r, N_COL2, fill=HEADER_FILL, font=HEADER_FONT_W)

    # -- 数据行 --
    total_d_run, total_d_stop, total_d_km = 0, 0, 0
    total_u_run, total_u_stop, total_u_km = 0, 0, 0
    n_down, n_up = 0, 0

    def t_str(val):
        if val is None: return "—"
        return f"{int(val)//60:02d}:{int(val)%60:02d}"

    def vd(val):
        return val if val is not None else "—"

    DATA_START = 7
    for i, row in enumerate(table2_rows):
        r = DATA_START + i
        data = {
            1: vd(row["down_id"]), 2: t_str(row["down_depart_a"]), 3: t_str(row["down_arrive_b"]),
            4: vd(row["down_travel"]), 5: vd(row["down_run"]), 6: vd(row["down_stop"]), 7: vd(row["down_km"]),
            8: vd(row["up_id"]), 9: t_str(row["up_depart_b"]), 10: t_str(row["up_arrive_a"]),
            11: vd(row["up_travel"]), 12: vd(row["up_run"]), 13: vd(row["up_stop"]), 14: vd(row["up_km"]),
            15: vd(row["loco_b_stay"]), 16: vd(row["loco_tow"]),
            17: t_str(row["loco_a_depart"]), 18: vd(row["loco_a_stay"]),
        }
        for col, val in data.items():
            ws2.cell(row=r, column=col, value=val)
        style_data_row(ws2, r, N_COL2)

        if row["down_run"]:
            total_d_run += row["down_run"]; total_d_stop += row["down_stop"]; total_d_km += row["down_km"]; n_down += 1
        if row["up_run"]:
            total_u_run += row["up_run"]; total_u_stop += row["up_stop"]; total_u_km += row["up_km"]; n_up += 1

    # 合计行
    r = DATA_START + len(table2_rows)
    total_data = {
        1: "合计", 2: "—", 3: "—", 4: total_d_run + total_d_stop, 5: total_d_run, 6: total_d_stop, 7: total_d_km,
        8: "合计", 9: "—", 10: "—", 11: total_u_run + total_u_stop, 12: total_u_run, 13: total_u_stop, 14: total_u_km,
        15: "—", 16: "—", 17: "—", 18: "—",
    }
    for col, val in total_data.items():
        ws2.cell(row=r, column=col, value=val)
    style_data_row(ws2, r, N_COL2)
    for col in range(1, N_COL2 + 1):
        ws2.cell(row=r, column=col).font = Font(name="微软雅黑", bold=True, size=10)

    # 注释
    r += 1
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N_COL2)
    ws2.cell(row=r, column=1, value="注：每一行上行方向车次按机车折返交路的顺序填写").font = Font(name="微软雅黑", size=9, italic=True)

    auto_width(ws2, N_COL2, min_width=6, max_width=16)

    # 保存
    filename = "附表1和2_计算结果.xlsx"
    try:
        wb.save(filename)
    except PermissionError:
        import time
        filename = f"附表1和2_计算结果_{int(time.time())}.xlsx"
        wb.save(filename)
    print(f"  [OK] Excel文件已保存到: {filename}")
    return filename


# ================================================================
# 第四部分：主程序
# ================================================================

class Tee:
    """同时输出到控制台和文件"""
    def __init__(self, filepath):
        self.file = open(filepath, "w", encoding="utf-8")
        self.stdout = sys.stdout

    def write(self, data):
        self.stdout.write(data)
        self.file.write(data)

    def flush(self):
        self.stdout.flush()
        self.file.flush()

    def close(self):
        self.file.close()


def main():
    # 输出文件
    output_file = "附表1和2_计算结果.txt"
    tee = Tee(output_file)
    sys.stdout = tee

    print("=" * 70)
    print("  A-B 区段运行图指标计算 — 附表1 & 附表2")
    print("=" * 70)

    # ---- 先算平行/非平行通过能力 ----
    print("\n\n第一部分：通过能力计算\n")
    sections = calc_restrictive_section()
    available_time = 1440 - T_MAINTENANCE
    restrictive = max(sections, key=lambda s: s["t_period"])
    n_parallel = available_time / restrictive["t_period"]
    n_parallel_int = int(n_parallel)
    parallel_result = {
        "available_time": available_time,
        "restrictive_section": restrictive["section"],
        "restrictive_period": restrictive["t_period"],
        "n_parallel_float": n_parallel,
        "n_parallel": n_parallel_int,
        "section_details": sections,
    }
    non_parallel_result = {
        "n_parallel": n_parallel_int,
        "n_non_parallel_float": n_parallel_int - 7.7,
        "n_non_parallel": int(n_parallel_int - 7.7),
    }

    # ---- 打印附表1 ----
    print_table1(parallel_result, non_parallel_result)

    # ---- 打印附表2 ----
    print("\n\n第二部分：质量指标计算\n")
    table2_rows = calc_table2_data()
    print_table2(table2_rows)

    # ---- 输出 Excel ----
    print("\n\n第三部分：生成Excel文件\n")
    xlsx_file = save_to_xlsx(parallel_result, non_parallel_result, table2_rows)

    print("\n" + "=" * 160)
    print("  计算完成！")
    print("=" * 160)

    # 恢复stdout并关闭文件
    sys.stdout = tee.stdout
    tee.close()
    print(f"\n  [OK] 结果已保存到: {output_file}")


if __name__ == "__main__":
    main()
